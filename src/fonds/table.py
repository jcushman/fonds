"""Reading and writing the inventory table.

The inventory is meant to be edited by hand as well as regenerated, so writing
it is a merge, not an overwrite: columns you added, notes you wrote above the
table, and rows for repos the tool cannot see are all preserved. That property
is why the inventory can double as the input to `fonds tags sync`.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime, timezone
from pathlib import Path

import click

NAME_COLUMN = "Name"


def order_fields(rows: list[dict], columns: list[str]) -> list[str]:
    """Known columns first, then any extras in order of first appearance."""
    seen = set(columns)
    extras = []
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                extras.append(key)
    return list(columns) + extras


class Format:
    ext: str

    def write(self, rows: list[dict], path: Path, columns: list[str]) -> None:
        raise NotImplementedError

    def read(self, path: Path) -> list[dict]:
        raise NotImplementedError

    def write_into(self, path: Path, rows: list[dict], columns: list[str]) -> None:
        """Replace just the data portion of an existing file."""
        self.write(rows, path, columns)

    def update(
        self, path: Path, rows: list[dict], columns: list[str], prune: bool
    ) -> list[str]:
        """Merge *rows* into an existing file. Returns names of stale repos."""
        existing = {row[NAME_COLUMN]: row for row in self.read(path)}
        incoming = {row[NAME_COLUMN]: row for row in rows}
        stale = [name for name in existing if name not in incoming]
        for name, row in incoming.items():
            existing.setdefault(name, {}).update(row)
        if prune:
            for name in stale:
                del existing[name]
        self.write_into(path, list(existing.values()), columns)
        return stale


class CsvFormat(Format):
    ext = ".csv"

    def write(self, rows: list[dict], path: Path, columns: list[str]) -> None:
        fields = order_fields(rows, columns)
        with path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)

    def read(self, path: Path) -> list[dict]:
        with path.open(newline="") as f:
            return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

_TABLE_LINE = re.compile(r"^\|.+\|$")
_SEP_LINE = re.compile(r"^\|[ :|-]+\|$")


def _escape(value) -> str:
    return str(value).replace("|", "\\|")


def _unescape(value: str) -> str:
    return value.replace("\\|", "|")


def render_table(rows: list[dict], columns: list[str]) -> str:
    fields = order_fields(rows, columns)
    lines = [
        "| " + " | ".join(fields) + " |",
        "| " + " | ".join("---" for _ in fields) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_escape(row.get(f, "")) for f in fields) + " |")
    return "\n".join(lines) + "\n"


def find_table(lines: list[str]) -> tuple[int, int] | None:
    """(start, end) line indices of the first markdown table, or None."""
    for i in range(len(lines) - 1):
        if _TABLE_LINE.match(lines[i]) and _SEP_LINE.match(lines[i + 1]):
            end = i + 2
            while end < len(lines) and _TABLE_LINE.match(lines[end]):
                end += 1
            return i, end
    return None


def _split_row(line: str, num_cols: int | None = None) -> list[str]:
    """Split a markdown row into cells, tolerating unescaped pipes.

    Overflow is folded back into the second cell (the description), which is
    where stray pipes realistically come from.
    """
    parts = line.strip("|").split("|")
    if num_cols is not None and len(parts) > num_cols:
        overflow = len(parts) - num_cols
        parts = parts[:1] + ["|".join(parts[1 : 2 + overflow])] + parts[2 + overflow :]
    return parts


def parse_table(lines: list[str], start: int, end: int) -> list[dict]:
    headers = [h.strip() for h in _split_row(lines[start])]
    rows = []
    for line in lines[start + 2 : end]:
        values = [_unescape(v.strip()) for v in _split_row(line, len(headers))]
        rows.append(dict(zip(headers, values)))
    return rows


class MarkdownFormat(Format):
    ext = ".md"

    def write(self, rows: list[dict], path: Path, columns: list[str]) -> None:
        path.write_text(render_table(rows, columns))

    def read(self, path: Path) -> list[dict]:
        lines = path.read_text().splitlines()
        span = find_table(lines)
        return parse_table(lines, *span) if span else []

    def write_into(self, path: Path, rows: list[dict], columns: list[str]) -> None:
        text = path.read_text()
        lines = text.splitlines(keepends=True)
        span = find_table([line.rstrip("\n") for line in lines])
        if span is None:
            path.write_text(text + "\n" + render_table(rows, columns))
            return
        start, end = span
        path.write_text(
            "".join(lines[:start]) + render_table(rows, columns) + "".join(lines[end:])
        )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

class ExcelFormat(Format):
    ext = ".xlsx"

    @staticmethod
    def _openpyxl():
        try:
            import openpyxl
        except ImportError:
            raise click.ClickException(
                "Excel output requires openpyxl. Install it with: "
                "uv add 'fonds[excel]' (or pip install openpyxl)"
            ) from None
        return openpyxl

    def read(self, path: Path) -> list[dict]:
        workbook = self._openpyxl().load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                break
            rows.append(dict(zip(headers, row)))
        return rows

    def write(self, rows: list[dict], path: Path, columns: list[str]) -> None:
        openpyxl = self._openpyxl()
        from openpyxl.styles import Alignment, Font, PatternFill

        from .plugins.inventory import column_style

        fields = order_fields(rows, columns)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Repos"

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="2D333B")
        header_align = Alignment(horizontal="center", vertical="center")
        cell_font = Font(name="Arial", size=10)
        center = Alignment(horizontal="center", vertical="center")
        wrap = Alignment(vertical="center", wrap_text=True)
        stripe = PatternFill("solid", start_color="F6F8FA")

        styles = {field: column_style(field) for field in fields}

        for index, title in enumerate(fields, start=1):
            cell = sheet.cell(row=1, column=index, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            sheet.column_dimensions[cell.column_letter].width = styles[title][0]

        sheet.row_dimensions[1].height = 22

        for row_index, row in enumerate(rows, start=2):
            for col_index, field in enumerate(fields, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=row.get(field, ""))
                cell.font = cell_font
                cell.alignment = center if styles[field][1] else wrap
                if row_index % 2 == 0:
                    cell.fill = stripe

        sheet.freeze_panes = "A2"
        last_column = sheet.cell(row=1, column=len(fields)).column_letter
        sheet.auto_filter.ref = f"A1:{last_column}1"

        footer = len(rows) + 3
        sheet.cell(row=footer, column=1, value=f"Total repos: {len(rows)}").font = Font(
            name="Arial", bold=True, size=10
        )
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        sheet.cell(row=footer, column=3, value=f"Generated: {stamp}").font = Font(
            name="Arial", italic=True, size=9, color="888888"
        )
        workbook.save(path)


FORMATS: dict[str, Format] = {
    "markdown": MarkdownFormat(),
    "csv": CsvFormat(),
    "excel": ExcelFormat(),
}

BY_EXTENSION: dict[str, Format] = {fmt.ext: fmt for fmt in FORMATS.values()}


def read_table(path: Path) -> list[dict]:
    """Read a table file, picking the format from its extension."""
    fmt = BY_EXTENSION.get(path.suffix)
    if fmt is None:
        raise click.ClickException(
            f"Unsupported file extension {path.suffix!r}. "
            f"Supported: {', '.join(sorted(BY_EXTENSION))}"
        )
    return fmt.read(path)


def row_tags(row: dict) -> list[str]:
    return [tag.strip() for tag in (row.get("Tags") or "").split(",") if tag.strip()]
